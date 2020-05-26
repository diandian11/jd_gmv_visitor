
# coding: utf-8

# In[2]:


import pandas as pd 
import numpy as np
from scipy.optimize import curve_fit  
import matplotlib.pyplot as plt 
import datetime
# from sklearn import cross_validation
from impala.dbapi import connect
from impala.util import as_pandas
import json
import pymysql
from sqlalchemy import create_engine
import os
import xlsxwriter
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
# 邮件模块
import smtplib
from smtplib import SMTP
from email.mime.text import MIMEText
from email.header import Header
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart 
import logging


# In[22]:


path1 =os.getcwd()
print("路径是在:",path1)
df_c = pd.read_excel(path1+'/'+'百丽官方旗舰店gmv转化系数.xlsx', sheet_name = '系数')
df_c.index= df_c['成交金额指数']


# In[25]:


def fund(x, a, b):  
    """gmv = a*exp(ln指数*b)"""
    return a * np.exp(b * x)


# In[26]:


def fun_fit(x):
    global df_c
    list3 = df_c.index
    if list3[0]<=x<list3[1]:
        a = df_c.loc[list3[0],'a']
        b = df_c.loc[list3[0],'b']
    elif list3[1]<=x<list3[2]:
        a = df_c.loc[list3[1],'a']
        b = df_c.loc[list3[1],'b']  
    elif list3[2]<=x<list3[3]:
        a = df_c.loc[list3[2],'a']
        b = df_c.loc[list3[2],'b']  
    elif list3[3]<=x<list3[4]:
        a = df_c.loc[list3[3],'a']
        b = df_c.loc[list3[3],'b']
    elif list3[4]<=x<list3[5]:
        a = df_c.loc[list3[4],'a']
        b = df_c.loc[list3[4],'b']
    elif list3[5]<=x<list3[6]:
        a = df_c.loc[list3[5],'a']
        b = df_c.loc[list3[5],'b']  
    elif list3[6]<=x<list3[7]:
        a = df_c.loc[list3[6],'a']
        b = df_c.loc[list3[6],'b']  
    elif list3[7]<=x<list3[8]:
        a = df_c.loc[list3[7],'a']
        b = df_c.loc[list3[7],'b']
    elif list3[8]<=x<list3[9]:
        a = df_c.loc[list3[8],'a']
        b = df_c.loc[list3[8],'b']  
    elif list3[9]<=x<list3[10]:
        a = df_c.loc[list3[9],'a']
        b = df_c.loc[list3[9],'b']  
    elif list3[10]<=x<list3[11]:
        a = df_c.loc[list3[10],'a']
        b = df_c.loc[list3[10],'b']
    elif list3[11]<=x<list3[12]:
        a = df_c.loc[list3[11],'a']
        b = df_c.loc[list3[11],'b']
    elif list3[12]<=x<list3[13]:
        a = df_c.loc[list3[12],'a']
        b = df_c.loc[list3[12],'b']  
    elif list3[13]<=x<list3[14]:
        a = df_c.loc[list3[13],'a']
        b = df_c.loc[list3[13],'b']  
    elif list3[14]<=x<list3[15]:
        a = df_c.loc[list3[14],'a']
        b = df_c.loc[list3[14],'b']
    elif list3[15]<=x<list3[16]:
        a = df_c.loc[list3[15],'a']
        b = df_c.loc[list3[15],'b']  
    elif list3[16]<=x<list3[17]:
        a = df_c.loc[list3[16],'a']
        b = df_c.loc[list3[16],'b']  
    elif list3[17]<=x<list3[18]:
        a = df_c.loc[list3[17],'a']
        b = df_c.loc[list3[17],'b']
    elif list3[18]<=x<list3[19]:
        a = df_c.loc[list3[18],'a']
        b = df_c.loc[list3[18],'b']
    elif list3[19]<=x<list3[20]:
        a = df_c.loc[list3[19],'a']
        b = df_c.loc[list3[19],'b']
    elif list3[20]<=x<list3[21]:
        a = df_c.loc[list3[20],'a']
        b = df_c.loc[list3[20],'b']
    elif list3[21]<=x<list3[22]:
        a = df_c.loc[list3[21],'a']
        b = df_c.loc[list3[21],'b']
    elif list3[22]<=x<list3[23]:
        a = df_c.loc[list3[22],'a']
        b = df_c.loc[list3[22],'b']
    elif list3[23]<=x<list3[24]:
        a = df_c.loc[list3[23],'a']
        b = df_c.loc[list3[23],'b']
    elif list3[24]<=x<list3[25]:
        a = df_c.loc[list3[24],'a']
        b = df_c.loc[list3[24],'b']
    elif list3[25]<=x<list3[26]:
        a = df_c.loc[list3[25],'a']
        b = df_c.loc[list3[25],'b']
    elif list3[26]<=x<list3[27]:
        a = df_c.loc[list3[26],'a']
        b = df_c.loc[list3[26],'b']
    elif list3[27]<=x<list3[28]:
        a = df_c.loc[list3[27],'a']
        b = df_c.loc[list3[27],'b']
    elif list3[28]<=x<list3[29]:
        a = df_c.loc[list3[28],'a']
        b = df_c.loc[list3[28],'b']
    else:
        a = df_c.loc[list3[29],'a']
        b = df_c.loc[list3[29],'b']
    y = fund(np.log(x),a,b)
    return y


# In[27]:


# fun_fit(4445555)


# In[28]:


#### 京东指数取数
# 连接集群hive/impala
def impala_connect(sql, **kwargs):
# impala 
    host = kwargs.get("host", 'impala.bjds.belle.lan')
    port = kwargs.get("port", 21051)    
    timeout = kwargs.get("timeout", 3600)
# hive
    # host = kwargs.get("host", 'impala.bjds.belle.lan')
    # port = kwargs.get("port", 10008)    
    # timeout = kwargs.get("timeout", 3600)
    user = kwargs.get("user", "lv.d.sz")
    password = kwargs.get("password", 'JHjLXpyQ')
    kerberos_service_name = kwargs.get("kerberos_service_name", "impala")
    conn = connect(host=host, port=port, timeout=timeout, user=user, password=password, kerberos_service_name=
                   kerberos_service_name,auth_mechanism='LDAP')
    cur = conn.cursor(user=user)
    if sql is not None:
        cur.execute(sql)
        try:
            df = as_pandas(cur)
        except:
            return cur
    return df

##mysql
# mysql 数据库连接
class MySQLConn(object):
    """Class supply connection to mysql database"""
    def __init__(self, **info):
        self.host = info.get("host", "10.251.36.8")
        self.user = info.get("user", "belle_read")
        self.port = info.get("port", 3306)
        self.password = info.get("password", "Belle@123456")
        self.db = info.get("db", "belle")
        self.table = info.get("table", "table_check")
        # self.table_add = info.get("table_add", None)
        self.charset = info.get("charset", "utf8")
        try:
            self.conn = pymysql.connect(host=self.host, port=self.port,
                                    user=self.user, password=self.password, db=self.db, charset=self.charset)
            self.cur = self.conn.cursor()
        except pymysql.err.OperationalError as e:
            print('Error is ' + str(e))
            sys.exit()

    def read_mysql(self, col=None, where_append=None, chunk_size=500000):
        if col is None:
            col = "*"
        if where_append is None:
            where_append = ''
        try:
            sql = 'select %s from %s %s' % (col, self.table, where_append)
            dfs = pd.read_sql(sql, con=self.conn, chunksize=chunk_size)
        except pymysql.err.ProgrammingError as e:
            print('Error is ' + str(e))
            sys.exit()
        dfs = list(dfs)
        if len(list(dfs)) == 0:
            return None
        else:
            return pd.concat(dfs)

    def to_sql(self, df, apply_date, *args):
        pass

    def close(self):
        self.conn.close()


# In[29]:


def act_node(date_str):
    if date_str<='2020-05-31':
        c = '预售期&预热期'
    elif date_str<='2020-06-15':
        c = '专场期'
    elif date_str<='2020-06-20':
        c = '高潮期'
    elif date_str == '2020-06-21':
        c = '还场期'
    return c


# In[30]:


def adj(df_check):
    """大小顺序调整"""
    for i in range(100):
        if df_check['排名'][i] !=df_check['gmv_rank'][i]:
            z= (np.log(df_check.loc[i,'成交金额指数'])-np.log(df_check.loc[i+1,'成交金额指数']))/(np.log(df_check.loc[i-1,'成交金额指数'])-np.log(df_check.loc[i+1,'成交金额指数']))                                                                                                                                                 
            df_check.loc[i,'成交金额']=   df_check.loc[i+1,'成交金额']+z*(df_check.loc[i-1,'成交金额']-df_check.loc[i+1,'成交金额'])
        df_check['gmv_rank'] = df_check['成交金额'].rank(method = 'first', ascending = False)
    return df_check

def data_check(df1_2, kind_list):
    """对dataframe 进行检验调整"""
    df3 = pd.DataFrame()
    for kind in kind_list:
        df_check = df1_2[df1_2['类目']== kind].reset_index(drop = True)
        df_check['gmv_rank'] = round(df_check['成交金额'].rank(method = 'first', ascending = False),0)
        df3_temp = adj(df_check)
        df3 = pd.concat([df3, df3_temp], axis =0)
    df3.drop(['gmv_rank'],axis =1, inplace = True)
    return df3


# In[31]:


## 写入数据到数据库，读取累计数据
def leiji_report(df2, date1):
    """写入一天的数据到数据库保留，并读出最近60天访客数"""
    self1 = {
    "host": "10.251.11.5",
    "user": "sz_analy",
    "port": 3306,
    "password": "Belle2020@sz",
    "db":"bi_analysis_sz",
    "table": "jd_618_daily_rank_temp"
    }
    mysql1 =  MySQLConn(**self1)# 改变连接的信息
    check = mysql1.read_mysql(col = 'max(`日期`) as dt, count(1) as count_n')#.iloc[0,0]
    dt_max = check.iloc[0,0]
    count_n = check.iloc[0,1]
    if  dt_max is None or date1 > dt_max:# 
        # 写入到mysql数据库
        conn = create_engine('mysql+mysqlconnector://sz_analy:Belle2020@sz@10.251.11.5:3306/bi_analysis_sz?charset=utf8',encoding ='utf-8')   
        pd.io.sql.to_sql(df2,'jd_618_daily_rank_temp', con=conn,chunksize=1000, schema='bi_analysis_sz',index=False,if_exists='append') # append 是追加的形式写入；replace
        print('品牌日访客数写入【%s日】数据库成功！共%s行'%(date1,len(df2)))
    else:
        print('数据已是最新，不重复写入！如果有问题删除[%s日】后重新刷入'%dt_max)
    # 取出近60天的访客数
    where_append =  "where datediff ('{}',`日期`) between 0 and 60".format(date1)
    df3 =  mysql1.read_mysql(where_append = where_append) # 取数近60天
    return df3


# In[32]:


sql_vender_day= """
-- 二级品类店铺top100日度数据
select category as `一级类目`,
child_category as `子类目`, 
start_day as `日期`, 
end_day,
date_type, 
rank  as `排名`, 
shop_name `店铺名`,
gmv_amount_index as `成交金额指数`,
order_num_index as `成交单量指数`,
follow_num as `关注人数`,
visitor_index as `访客指数`,
search_click_index as `搜索点击指数`
from  bi_analysis.jd_sz_vender_ranks where 
category in ('时尚女鞋','流行男鞋','潮流女包')
 and child_category = ""
and start_day = '%s'
and end_day = '%s'
-- and length(date_type) = 6
and belle_shop_name = '百丽官方旗舰店'
order by start_day, category, rank
"""


# In[33]:


def write2excel(df,sh,date):
    border_set = Border(left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin',color='000000'),
                            top=Side(border_style='thin',color='000000'),
                            bottom=Side(border_style='thin',color='000000')) ## 单元格边框
    col_zip = list(df.columns)
    for j, col0 in enumerate(col_zip):
        col1 = list(df[col0])
        for i, item in enumerate(col1):
            c1 = sh.cell(row = 5+i,column=j+1,value=item)
            c1.border = border_set 
    sh.cell(row =3 ,column =1, value = date)


# In[34]:


def jd2excel(file_path,df3,kind_list,last_day):
    df = df3
    wb = load_workbook(file_path+'/'+'京东618Top100竞店数据模板v1.xlsx')
    for kind in kind_list:
        sh1 = wb[kind]
        df1 = df[df['类目']==kind]
        df1.sort_values(by =['日期','排名'], ascending = [False, True],inplace =True)
        write2excel(df = df1,sh=sh1,date = last_day)
    wb.save(file_path+'/result/'+'京东618Top100竞店数据【%s】.xlsx'% last_day)
    wb.close()


# In[35]:


### 发送邮件
# 正常邮件推送
def send_mail(flag,date1, filename, file_to_path):
    if flag == 1:
        text_body = """大家好，\n    【%s】618京东Top100竞店排行榜数据如附件所示, 请查收!\n\n
    附件说明：1.excel包含 时尚女鞋、流行男鞋、潮流女包3个类目Top100榜单；
             2. 数据按日期顺序累计存放。
        """%(date1)
    else:
        text_body = "None"
    sender = 'lv.d.sz@belle.com.cn'
    receivers = ['lv.d.sz@belle.com.cn']
#     receivers = [
#         'zhou.x.sz@belle.com.cn',
#         'huang.hd.sz@belle.com.cn',
#         'yu.hr.sz@belle.com.cn',
#         'li.qy.sz@belle.com.cn'
#     ]  # 接收邮件，可设置为你的QQ邮箱或者其他邮箱
#     CC = [ 'xiang.d@belle.com.cn',
#         'zheng.q.sz@belle.com.cn',
#         'he.xh.sz@belle.com.cn',
#         'lv.d.sz@belle.com.cn'
#           ]# 抄送
    CC = [ 'lv.d.sz@belle.com.cn']# 抄送

    # 三个参数：第一个为文本内容，第二个 plain 设置文本格式，第三个 utf-8 设置编码
    message = MIMEMultipart()
    message['From'] = Header('分析组 <10.240.20.22>', charset='utf-8')
    message['To'] = ";".join(receivers)
    message['CC'] = ";".join(CC)

    subject = '京东618Top100竞店日度数据【%s】'%(date1)
    message['Subject'] = Header(subject, charset='utf-8')

    # 邮件正文
    message.attach(MIMEText(text_body, 'plain', 'utf-8'))

    # 构造附件

    att1 = MIMEText(open(file_to_path+'/result/'+filename+'.xlsx', 'rb').read(), 'base64', 'utf-8')
    att1["Content-Type"] = 'application/octet-stream'
    # att1["Content-Disposition"] = 'attachment; filename="%s"' % filename_show.encode('utf-8')
    att1.add_header('Content-Disposition', 'attachment', filename=Header(filename+'.xlsx', 'utf-8').encode())
    message.attach(att1)

    try:
        smtpObj = smtplib.SMTP_SSL('smtp.exmail.qq.com', port=465)
        # smtpObj.connect()
        smtpObj.login('lv.d.sz@belle.com.cn',password='ld971646OK')
        smtpObj.sendmail(sender, receivers + CC, message.as_string())
        print("邮件发送成功")
    except smtplib.SMTPException:
        print(smtplib.SMTPException.errno)


# In[36]:


### 发送邮件
# 异常邮件推送
def send_mail2(flag,date1,n):
    if flag == 1:
        text_body = """数据未更新异常！   【%s】618京东Top100竞店排行榜数据未更新，请检查。\n\n
        数据只有【%d】条
        """%(date1,n)
    else:
        text_body = "None"
    sender = 'lv.d.sz@belle.com.cn'
    # receivers = ['lv.d.sz@belle.com.cn']
    receivers = [
        'lv.d.sz@belle.com.cn'
    ]  # 接收邮件，可设置为你的QQ邮箱或者其他邮箱
    CC = [

        'lv.d.sz@belle.com.cn'
          ]# 抄送
    # CC = [ 'lv.d.sz@belle.com.cn']# 抄送         'he.xh.sz@belle.com.cn',

    # 三个参数：第一个为文本内容，第二个 plain 设置文本格式，第三个 utf-8 设置编码
    message = MIMEMultipart()
    message['From'] = Header('分析组 <10.240.20.22>', charset='utf-8')
    message['To'] = ";".join(receivers)
    message['CC'] = ";".join(CC)

    subject = '异常！京东618竞店日度数据【%s】'%(date1)
    message['Subject'] = Header(subject, charset='utf-8')

    # 邮件正文
    message.attach(MIMEText(text_body, 'plain', 'utf-8'))

    try:
        smtpObj = smtplib.SMTP_SSL('smtp.exmail.qq.com', port=465)
        # smtpObj.connect()
        smtpObj.login('lv.d.sz@belle.com.cn',password='ld971646OK')
        smtpObj.sendmail(sender, receivers + CC, message.as_string())
        print("邮件发送成功")
    except smtplib.SMTPException:
        print(smtplib.SMTPException.errno)


# In[37]:


if __name__ == '__main__':
    last_day = datetime.date.strftime(datetime.date.today()+datetime.timedelta(-1),"%F")
#     last_day ='2020-05-24'
    sql_f = sql_vender_day%(last_day,last_day)
    df1 = impala_connect(sql_f)
    # 数据检查
    if len(df1)==300:
        df1_1 =df1.drop(['子类目','end_day','date_type'],axis =1)
        df1_1['渠道'] = '京东'
        df1_1['活动节点'] = act_node(last_day)
        df1_1['成交金额'] = df1_1['成交金额指数'].map(lambda x:fun_fit(x))
        df1_1['访客数'] = None                        #df1_1['访客指数'].map(lambda x:fun_fit(x))
        # df1_1['点击次数'] = '-'                        #df1_1['搜索点击指数'].map(lambda x:fun_fit(x))
        col_name ={'一级类目':'类目','店铺名':'商家名称'}
        df1_1.rename(columns = col_name, inplace = True)
        df1_2 =df1_1[['活动节点','渠道','类目','排名','日期','商家名称','成交金额','成交金额指数','关注人数','成交单量指数','访客数','访客指数','搜索点击指数']]
        kind_list = ['时尚女鞋','流行男鞋','潮流女包']
        df_r = data_check(df1_2, kind_list)
        df3 = leiji_report(df2 =df_r , date1 = last_day)
        jd2excel(file_path = path1,df3 =df3,kind_list = kind_list,last_day =last_day) # 写入数据到excel
        send_mail(flag=1,date1= last_day,file_to_path =path1, filename = '京东618Top100竞店数据【%s】'%last_day)
    else:
        send_mail2(flag=1,date1=last_day,n =len(df1))

